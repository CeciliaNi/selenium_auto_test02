import openpyxl
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import matplotlib.pyplot as plt
from openpyxl.styles import Alignment


# filename = r'c:\Users\wangwenbo\selenium_auto_test01\test_suite.xls'
# excel = win32.Dispatch('Excel.Application')
#
# wb = excel.Workbooks.Open(filename)
# wb.SaveAs(filename + "x", FileFormat=51)
#
# wb.Close()
# excel.Application.Quit()

align = Alignment(horizontal='distributed', vertical='center', wrap_text=True)


def write_into_excel(resultwb, filename, sheet_name, row, column, value):
    """
    往excel表格中写入测试结果数据
    :return:
    """
    # resultwb = load_workbook(filename)

    ws = resultwb[sheet_name]
    ws.cell(row=row, column=column, value=value).alignment = align
    resultwb.save(filename)


def img_into_excel(resultwb, sheet_name, row, column, img_path):
    """
    往excel中插入图片
    :return:
    """

    # resultwb = load_workbook(filename)
    ws = resultwb[sheet_name]

    # 设置文字图片单元格的行高列宽
    column_width = 12.25
    row_height = 80.10

    ws.column_dimensions[column].width = column_width  # 修改列的列宽
    ws.row_dimensions[row].height = row_height  # 修改行的行高

    # 插入图片的单元格
    target_cell = column+str(row)
    if os.path.exists(img_path):
        img = Image(img_path)
        # 设置图片的宽高
        newsize = (90, 90)
        img.width, img.height = newsize

        # 插入图片
        ws.add_image(img, target_cell)

        # 保存
        # resultwb.save(filename)


def pie_report(filename):
    """

    :return:
    """
    # 解决中文显示问题
    plt.rcParams['font.sans-serif'] = ['SimHei']
    plt.rcParams['axes.unicode_minus'] = False

    resultwb = load_workbook(filename)
    ws = resultwb['TestSuite']

    succ_num = 0
    fail_num = 0
    for i, cell in enumerate(ws['G']):
        if i > 0:
            if cell.value == 'success':
                succ_num += 1
            elif cell.value == 'fail':
                fail_num += 1

    labels = ['success', 'fail']
    data = [succ_num, fail_num]

    plt.pie(data, labels=labels, autopct='%1.2f%%')  # 画饼图（数据，数据对应的标签，百分数保留两位小数点）
    plt.title("Pie chart")
    plt.savefig("../report/PieChart.png")
    plt.show()

    img = Image('../report/PieChart.png')
    ws.add_image(img, 'A11')
    resultwb.save(filename)

