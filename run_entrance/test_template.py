"""
作者：倪媛
功能：利用python+excel+unittest+ddt接口自动化数据驱动测试用例
日期：28/4/2019
"""
import unittest
import ddt
from shutil import copyfile
from openpyxl import load_workbook
import pandas as pd
import json
from run_entrance.keyword_def import key_word_func
from run_entrance.write_excel import write_into_excel
from picture.save_img import save_img
from run_entrance.write_excel import img_into_excel
from driver.driver import driver
import os

# 获取该测试文件的全称，通过一系列操作得到最后'_'后的名称 例如'test_api_login.py'--》login
full_filename = os.path.basename(__file__)
filename = full_filename[:full_filename.rfind('.')]
last_name = filename.split('_')[-1]

testxlsx = '../run_entrance/test_suite/test_suite_' + last_name + '.xlsx'
report_file = '../report/report_'+last_name+'.xlsx'

test_step_datas = pd.read_excel(testxlsx, sheet_name='TestSteps', converters={u'操作值': str})
page_element_datas = pd.read_excel(testxlsx, sheet_name='PageElements')
config_content = pd.read_excel(testxlsx, sheet_name='TestConfig')

url = config_content['url'].values[0]


class ExcelUtil:
    def __init__(self, excelpath, sheetname):
        self.wb = load_workbook(excelpath)
        self.ws = self.wb[sheetname]
        # 获取所有行的数据
        self.rowdata = self.ws.iter_rows()

    def dict_data(self):
        rows = []
        # 获取行数据入列表
        # row的类型为元组
        # 内容为(< Cell 'TestSuite'.A1 >, < Cell 'TestSuite'.B1 >, < Cell 'TestSuite'.C1 >, < Cell 'TestSuite'.D1 >)
        for row in self.rowdata:
            rows.append(row)

        if len(rows) <= 1:
            print("总行数小于1")
        else:
            # 测试用例序号列表
            r = []
            # 将文档中所有的测试用例入字典s,字典s入列表r
            for row_num, each_row in enumerate(rows):
                s = {}
                if row_num > 0:
                    s['rownum'] = row_num + 1
                    s['测试用例序号'] = each_row[0].value
                    r.append(s)
            return r


def process_suite(suite_no):
    """
    对每行的测试用例进行运行并返回测试结果
    :return:
    """
    rownum = suite_no['rownum']
    suite_value = suite_no['测试用例序号']
    # 过滤出属于该测试用例的测试步骤数据
    filter_step = test_step_datas[test_step_datas['测试用例序号'] == suite_value]
    test_steps = filter_step['测试步骤描述'].values

    # 测试步骤是否成功标志
    is_step_succ = True
    for each_step in test_steps:
        # 找到PageElements页中的相关操作参数
        filter_page_element = page_element_datas[page_element_datas['页面元素'] == each_step]
        key_word = filter_page_element['关键字'].values[0]
        type = filter_page_element['页面定位元素'].values[0]
        loc = filter_page_element['页面元素定位表达式'].values[0]

        filter_test_step = test_step_datas[test_step_datas['测试步骤描述'] == each_step]
        # 找到该测试步骤对应的行号
        row_num = filter_test_step.index.values[0]
        # 找到TestSteps中的操作值
        if filter_test_step['操作值'].isnull().values:
            value = None
        else:
            value = filter_test_step['操作值'].values[0]

        # 获取TestSteps中的检查点的值
        if filter_test_step['检查点'].isnull().values:
            checkpoint = None
        else:
            # 读取文件中的检查点数据类型是str, 通过json.loads()转换为dict
            checkpoint = json.loads(filter_test_step['检查点'].values[0])

        # 获取TestStep中等待时间的值
        if filter_test_step['等待时间'].isnull().values:
            sleep_time = None
        else:
            sleep_time = filter_test_step['等待时间'].values[0]

        # 调用页面操作函数
        is_step_succ = key_word_func(key_word, type, loc, checkpoint=checkpoint, value=value, sleep_time=sleep_time)

        if is_step_succ:
            # 一个测试步骤运行结束后 将TestSteps页中的'测试结果'列填充完整
            write_into_excel(report_file, 'TestSteps', row_num + 2, 7, '用例步骤执行成功')
        else:
            # 一个测试步骤运行结束后 将TestSteps页中的'测试结果'列填充完整
            write_into_excel(report_file, 'TestSteps', row_num + 2, 7, '用例步骤执行失败')
            # 测试步骤失败时 截图保存并将照片插入单元格中
            test_suite_name = last_name
            pic_path = save_img(test_suite_name)
            # 当测试步骤返回不成功时，不再执行剩下的步骤，直接执行下个测试用例
            break

    # 一个用例运行结束后 将TestSuite页中的'是否执行','执行结果'列填充完整
    if is_step_succ:
        write_into_excel(report_file, 'TestSuite', rownum, 6, 'yes')
        write_into_excel(report_file, 'TestSuite', rownum, 7, 'success')
    else:
        write_into_excel(report_file, 'TestSuite', rownum, 6, 'yes')
        write_into_excel(report_file, 'TestSuite', rownum, 7, 'fail')
        img_into_excel(report_file, 'TestSteps', row_num + 2, 'H', pic_path)

    return is_step_succ


# 得出结果为[{'rownum': 2, '测试用例序号': 'HKZW-001'}, {'rownum': 3, '测试用例序号': 'HKZW-002'}, {'rownum': 4, '测试用例序号': 'HKZW-003'}]
testdata = ExcelUtil(testxlsx, 'TestSuite').dict_data()


@ddt.ddt
class TestApi(unittest.TestCase):  # 继承unittest.TestCase
    # 必须使用@classmethod 装饰器,所有test运行前运行一次
    @classmethod
    def setUpClass(cls):
        # 为了不污染测试的数据，出报告的时候先将test_suite_.xlsx复制到report目录下的report.xlsx

        copyfile(testxlsx, report_file)
        driver.delete_all_cookies()
        driver.get(url)

    @ddt.data(*testdata)
    def test_api(self, data):
        suite_result = process_suite(data)
        print('{}返回实际结果->：{}'.format(data['测试用例序号'], suite_result))
        # 断言
        self.assertTrue(suite_result)


if __name__ == "__main__":
    unittest.main()
